import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_pivot_tables():
    # Create pivot tables directory if it doesn't exist
    os.makedirs('pivot_tables', exist_ok=True)
    
    # Load processed data
    sales_df = pd.read_csv('../data/processed/processed_sales.csv')
    customers_df = pd.read_csv('../data/processed/processed_customers.csv')
    products_df = pd.read_csv('../data/processed/processed_products.csv')
    
    # Create Excel workbook
    wb = Workbook()
    
    # 1. Sales Analysis by Region and Category
    ws_sales = wb.create_sheet("Sales Analysis")
    
    # Create pivot table data using pandas
    sales_pivot = pd.pivot_table(
        sales_df,
        values=['TotalPrice', 'Quantity'],
        index=['Region', 'ProductID'],
        columns=[],
        aggfunc={'TotalPrice': np.sum, 'Quantity': np.sum},
        margins=True
    ).reset_index()
    
    # Add data to worksheet
    for r in dataframe_to_rows(sales_pivot, index=False, header=True):
        ws_sales.append(r)
    
    # Format headers
    for cell in ws_sales[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # 2. Customer Analysis
    ws_customers = wb.create_sheet("Customer Analysis")
    
    # Create customer pivot table
    customers_pivot = pd.pivot_table(
        customers_df,
        values=['CustomerID'],
        index=['JoinDate'],
        aggfunc={'CustomerID': 'count'},
        margins=True
    ).reset_index()
    
    for r in dataframe_to_rows(customers_pivot, index=False, header=True):
        ws_customers.append(r)
    
    # Format headers
    for cell in ws_customers[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # 3. Product Analysis
    ws_products = wb.create_sheet("Product Analysis")
    
    # Create product pivot table
    product_pivot = pd.pivot_table(
        products_df,
        values=['Price', 'Cost'],
        index=['Category', 'ProductName'],
        aggfunc={'Price': np.mean, 'Cost': np.mean},
        margins=True
    ).reset_index()
    
    for r in dataframe_to_rows(product_pivot, index=False, header=True):
        ws_products.append(r)
    
    # Format headers
    for cell in ws_products[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Save workbook
    wb.save("pivot_tables/retail_pivot_tables.xlsx")
    print("Pivot tables created successfully!")

if __name__ == "__main__":
    create_pivot_tables()
