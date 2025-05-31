import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_excel_analysis():
    # Create output directory if it doesn't exist
    os.makedirs('analysis', exist_ok=True)
    
    # Load processed data
    sales_df = pd.read_csv('../data/processed/processed_sales.csv')
    customers_df = pd.read_csv('../data/processed/processed_customers.csv')
    products_df = pd.read_csv('../data/processed/processed_products.csv')
    segments_df = pd.read_csv('../data/processed/customer_segments.csv')
    
    # Create Excel workbook
    wb = Workbook()
    
    # Add summary sheet
    summary = wb.active
    summary.title = "Summary"
    
    # Add basic statistics
    summary['A1'] = "Sales Analysis Summary"
    summary['A1'].font = Font(bold=True, size=14)
    
    # Add key metrics
    metrics = [
        ('Total Sales', sales_df['TotalPrice'].sum()),
        ('Total Orders', sales_df['OrderID'].nunique()),
        ('Average Order Value', sales_df['TotalPrice'].mean()),
        ('Total Customers', customers_df['CustomerID'].nunique()),
        ('Total Products', products_df['ProductID'].nunique())
    ]
    
    for i, (metric, value) in enumerate(metrics, start=2):
        summary[f'A{i}'] = metric
        summary[f'B{i}'] = value
        summary[f'A{i}'].font = Font(bold=True)
    
    # Add Sales by Category
    sales_by_category = pd.merge(sales_df, products_df, on='ProductID')\
        .groupby('Category')['TotalPrice'].sum().reset_index()
    
    ws_category = wb.create_sheet("Sales by Category")
    for r in dataframe_to_rows(sales_by_category, index=False, header=True):
        ws_category.append(r)
    
    # Add chart
    chart = BarChart()
    chart.title = "Sales by Category"
    chart.style = 10
    chart.x_axis.title = 'Category'
    chart.y_axis.title = 'Sales Amount'
    
    data = Reference(ws_category, min_col=2, min_row=1, max_col=2, max_row=len(sales_by_category)+1)
    cats = Reference(ws_category, min_col=1, min_row=2, max_row=len(sales_by_category)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_category.add_chart(chart, "E3")
    
    # Add Customer Segmentation
    segment_summary = segments_df.groupby('Segment').agg({
        'CustomerID': 'count',
        'Recency': 'mean',
        'Frequency': 'mean',
        'Monetary': 'mean'
    }).reset_index()
    
    ws_segments = wb.create_sheet("Customer Segments")
    for r in dataframe_to_rows(segment_summary, index=False, header=True):
        ws_segments.append(r)
    
    # Add what-if analysis sheet
    what_if = wb.create_sheet("What-If Analysis")
    what_if['A1'] = "What-If Analysis"
    what_if['A1'].font = Font(bold=True, size=14)
    
    # Add discount impact analysis
    what_if['A3'] = "Discount Impact Analysis"
    what_if['A3'].font = Font(bold=True)
    
    # Create discount scenarios
    discount_scenarios = {
        'Current': 0,
        '5% Discount': 0.05,
        '10% Discount': 0.10,
        '15% Discount': 0.15,
        '20% Discount': 0.20
    }
    
    # Calculate impact on sales
    base_sales = sales_df['TotalPrice'].sum()
    row = 4
    for name, discount in discount_scenarios.items():
        what_if[f'A{row}'] = name
        what_if[f'B{row}'] = f"{discount*100}%"
        what_if[f'C{row}'] = base_sales * (1 - discount)
        what_if[f'D{row}'] = f"=C{row}-C3"
        row += 1
    
    # Add headers
    what_if['A4'] = "Scenario"
    what_if['B4'] = "Discount"
    what_if['C4'] = "Projected Sales"
    what_if['D4'] = "Change"
    
    # Format headers
    for cell in what_if[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Save workbook
    wb.save("analysis/retail_analysis.xlsx")
    print("Excel analysis file created successfully!")

if __name__ == "__main__":
    create_excel_analysis()
